# pip install XlsxWriter pandas xlsxwriter xlrd xlwt openpyxl
import requests
import pandas as pd
from functions import *
import openpyxl
from openpyxl.styles import (
                        PatternFill, Border, Side,
                        Alignment, Font, GradientFill
                        )
from openpyxl import Workbook

response = requests.get('https://api.marketinout.com/run/screen?key=905d08db5fa24f77')
text = response.text
content = response.content

with open("response_2.txt", "w", encoding='utf8') as f:
    f.write(response.text)  # Записываем ответ от сервера в текстовый файл
    print('1. Ответ сервера записан в response_2.txt')

with open("response_2.txt", "r", encoding='utf8') as f:
    with open("response_2_modify.txt", "w", encoding='utf8') as m:
        for line in f:
            if not line.isspace():
                m.write(line)
        print('2. response_2.txt преобразован в response_2_modify.txt')

df_1 = pd.DataFrame({})  # Создаем пустой датафрейм, чтобы добавлять к нему строки
count = 0  # Счетчик строк в датафрейме
count_pass = 0  # Счетчик пропущенных строк

with open('response_2_modify.txt') as file:
    for line in file:
        list_ = list_float(line.removesuffix('\n').split('|'))
        # Проверка на количество пустых строк в списке-----------------------------
        # Пропустить список, если таких строк больше count_none
        count_none = 0
        for i in list_:
            try:
                if len(i) == 0:
                     count_none += 1
            except(TypeError):
                pass
        if count_none > 8:
            count_pass += 1
            continue
        # --------------------------------------------------------------------------
        ticker =              list_[0].removesuffix('.HK')
        name =                0
        last =                list_[1]
        date =                list_[2]
        altman =              list_[3]
        debt_equity =         list_[4]
        interest_cover =      list_[5]
        current_ratio =       list_[6]
        piotroski =           list_[7]
        beneish =             list_[8]
        intrinsic_potential = list_[9]
        lynch_potential =     list_[10]
        graham_potential =    list_[11]
        shiller =             list_[12]
        roa =                 list_[13]
        roa_5y =              list_[14]
        roe =                 list_[15]
        roe_5y =              list_[16]
        dvd_yield_y =         list_[17]
        dvd_yield_5y_avg =    list_[18]
        avol =                list_[19]

        df_1_new_row = pd.DataFrame({
            'Тикер': [ticker],
            # 'Название': [name],
            'Цена': [last],
            'Дата': [date],
            # 'Ликвидность': [likv(avol*last)],
            'Честность': [chestn(beneish)],
            'Эффективность': [effect(piotroski)],
            'Устойчивость (по вероятности банкротства)': [ust_bankr(altman)],
            'Устойчивость (по долговой нагрузке)': [ust_dolg(debt_equity, interest_cover, current_ratio)],
            'Рентабельность активов (ROA ttm), %': [if_none(roa)],
            'Рентабельность активов (ROA) ср. за 5 лет, %': [if_none(roa_5y)],
            'Рентабельность собств. капитала (ROE ttm), %': [if_none(roe)],
            'Рентабельность собств. капитала (ROE) ср. за 5 лет, %': [if_none(roe_5y)],
            'P/E Шиллера (CAPE)': [if_none(shiller)],
            'Потенциал по Бенджамину Грэму, %': [if_none(graham_potential)],
            'Потенциал по Питеру Линчу, %': [if_none(lynch_potential)],
            'Потенциал по прогнозируемому FCF, %': [if_none(intrinsic_potential)],
            'Стратегия Акции Стоимости': [akc_stoim(chestn(beneish), ust_bankr(altman), graham_potential)],
            # 'Див. доходность, %': [dvd_yield_y],
            # 'Див. доходность ср. за 5 лет, %': [dvd_yield_5y_avg],

        })

        df_1 = pd.concat([df_1, df_1_new_row], ignore_index=True)  # сцепить датафреймы
        count += 1

print(f'3. Создано строк в датафрейме: {count}')
print(f'4. Пропущено строк: {count_pass}')

xl_name = '2023.XX.XX_china_stocks.xlsx'
# Указать writer библиотеки
writer = pd.ExcelWriter(xl_name, engine='xlsxwriter')
df_1.to_excel(writer, 'Sheet1', index=False)  # Записать ваш DataFrame в файл

writer.close()  # Сохраним результат
print('5. Датафрейм записан в файл .xlsx')

# Наводим красоту --------------------------------------------------------------------------------
wb = openpyxl.load_workbook(xl_name)  # читаем excel-файл
ws = wb['Sheet1']  # получаем лист, с которым будем работать
ws.row_dimensions[1].height = 75  # Высота первой строки в пикселах

# Задаем ширину столбцов НЕ в пикселах
ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 8
ws.column_dimensions['C'].width = 25
list_columns = [chr(i) for i in range(ord('D'), ord('Z'))]  # Последовательность от E до ...+1
for i in list_columns:
    ws.column_dimensions[i].width = 16

# Заливка и перенос текста во всём листе
for row in ws:
    for cell in row:
        cell.alignment = Alignment(wrapText=True, horizontal="center", vertical="center")  # Перенос текста
        if any([cell.value == 'Высокая', cell.value == '-', cell.value == 'Покупать']):
            cell.fill = PatternFill('solid', fgColor="32CD32")
        elif any([cell.value == 'Средняя', cell.value == 'Держать']):
            cell.fill = PatternFill('solid', fgColor="ffff66")
        elif any([cell.value == 'Низкая', cell.value == 'Продавать', cell.value == 'Неликвидные']):
            cell.fill = PatternFill('solid', fgColor="FA8072")

# Выравнивание влево для первых столбцов
list_columns = ['A', 'B']
for i in list_columns:
    for cell in ws[i]:
        cell.alignment = Alignment(wrapText=False, horizontal="left", vertical="center")

# Заливка столбцов, используя свою функцию
gr_yell_red(ws, 'No', 'H', 'K', 20, 10)  # Заливка столбцов с рентабельностью
gr_yell_red(ws, 'Yes', 'L', 'L', 15, 15)  # Заливка столбца с P/E Шиллера
gr_yell_red(ws, 'No', 'M', 'M', 40, 0)  # Заливка столбца Потенциал по Бенджамину Грэму, %
gr_yell_red(ws, 'No', 'N', 'N', 40, 0)  # Заливка столбца Потенциал по Питеру Линчу, %
gr_yell_red(ws, 'No', 'O', 'O', 40, 0)  # Заливка столбца Потенциал по прогнозируемому FCF, %

# Закрепляем области, которые выше и левее указанной ячейки
ws.freeze_panes = 'B2'

wb.save(xl_name)  # записываем файл
