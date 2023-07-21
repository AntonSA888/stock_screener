# pip install XlsxWriter pandas xlsxwriter xlrd xlwt openpyxl python-dotenv
import requests
import datetime
import pandas as pd
from functions import *
import openpyxl
import os
from dotenv import load_dotenv, find_dotenv


load_dotenv(find_dotenv())


def value_strategy(url, xl_name):
    response = requests.get(url)

    with open(f"responses/response_{xl_name}.txt", "w", encoding='utf8') as f:
        f.write(response.text)  # Записываем ответ от сервера в текстовый файл
        print(f'-----{xl_name}-----\n1. Ответ сервера записан в responses/response_...txt')

    with open(f"responses/response_{xl_name}.txt", "r", encoding='utf8') as f:
        with open(f"responses/response_modify_{xl_name}.txt", "w", encoding='utf8') as m:
            for line in f:
                if not line.isspace():
                    m.write(line)
            print('2. response_...txt преобразован в response_modify_...txt')

    df_1 = pd.DataFrame({})  # Создаем пустой датафрейм, чтобы добавлять к нему строки
    count = 0  # Счетчик строк в датафрейме
    count_pass = 0  # Счетчик пропущенных строк

    with open(f'responses/response_modify_{xl_name}.txt') as file:
        for line in file:
            list_ = list_float(line.removesuffix('\n').split('|'))
            # Проверка на количество пустых строк в списке-----------------------------
            # Пропустить список, если таких строк больше count_none
            count_none = 0
            for i in list_:
                try:
                    if len(i) == 0:
                        count_none += 1
                except(TypeError):
                    pass
            if count_none > 8:
                count_pass += 1
                continue

            # Присваиваем переменным значения из списка
            ticker = list_[0].removesuffix('.ME').removesuffix('.HK')
            last = list_[1]
            date = list_[2]
            altman = list_[3]
            debt_equity = list_[4]
            interest_cover = list_[5]
            current_ratio = list_[6]
            piotroski = list_[7]
            beneish = list_[8]
            intrinsic_potential = list_[9]
            lynch_potential = list_[10]
            graham_potential = list_[11]
            shiller = list_[12]
            roa = list_[13]
            roa_5y = list_[14]
            roe = list_[15]
            roe_5y = list_[16]
            dvd_yield_y = list_[17]
            dvd_yield_5y_avg = list_[18]
            avol = list_[19]
            # Поскольку ликвидность считаем в рублях, цену акции тоже приводим к рублю, умножая на курс
            if url == os.getenv('api_china_stocks'):
                last_rub = last * 12
            elif url == os.getenv('api_russian_stocks'):
                last_rub = last
            else:
                print('!!!!! last_rub не вычислен !!!!!')

            df_1_new_row = pd.DataFrame({
                'Тикер': [ticker],
                'Цена': [last],
                'Дата': [date],
                'Ликвидность': [likv(avol * last_rub)],
                'Честность': [chestn(beneish)],
                'Эффективность': [effect(piotroski)],
                'Устойчивость (по вероятности банкротства)': [ust_bankr(altman)],
                'Устойчивость (по долговой нагрузке)': [ust_dolg(debt_equity, interest_cover, current_ratio)],
                'Рентабельность активов (ROA ttm), %': [if_none(roa)],
                'Рентабельность активов (ROA) ср. за 5 лет, %': [if_none(roa_5y)],
                'Рентабельность собств. капитала (ROE ttm), %': [if_none(roe)],
                'Рентабельность собств. капитала (ROE) ср. за 5 лет, %': [if_none(roe_5y)],
                'P/E Шиллера (CAPE)': [if_none(shiller)],
                'Потенциал по Бенджамину Грэму, %': [if_none(graham_potential)],
                'Потенциал по Питеру Линчу, %': [if_none(lynch_potential)],
                'Потенциал по прогнозируемому FCF, %': [if_none(intrinsic_potential)],
                'Стратегия Акции Стоимости': [akc_stoim(chestn(beneish), ust_bankr(altman), graham_potential)],

            })

            df_1 = pd.concat([df_1, df_1_new_row], ignore_index=True)  # сцепить датафреймы
            count += 1

    print(f'3. Создано строк в датафрейме: {count}')
    print(f'4. Пропущено строк: {count_pass}')

    # Указать writer библиотеки
    writer = pd.ExcelWriter(f'spreadsheets/{xl_name}', engine='xlsxwriter')
    df_1.to_excel(writer, 'Sheet1', index=False)  # Записать ваш DataFrame в файл

    writer.close()  # Сохраним результат
    print(f'5. Датафрейм записан в spreadsheets/...xlsx\n')

    # Наводим красоту --------------------------------------------------------------------------------
    wb = openpyxl.load_workbook(f'spreadsheets/{xl_name}')  # читаем excel-файл
    ws = wb['Sheet1']  # получаем лист, с которым будем работать
    ws.row_dimensions[1].height = 75  # Высота первой строки в пикселах

    # Задаем ширину столбцов НЕ в пикселах
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 25
    list_columns = [chr(i) for i in range(ord('D'), ord('Z'))]  # Последовательность от E до ...+1
    for i in list_columns:
        ws.column_dimensions[i].width = 16

    # Заливка и перенос текста во всём листе
    for row in ws:
        for cell in row:
            cell.alignment = Alignment(wrapText=True, horizontal="center", vertical="center")  # Перенос текста
            if any([cell.value == 'Высокая', cell.value == '-', cell.value == 'Покупать']):
                cell.fill = PatternFill('solid', fgColor="32CD32")
            elif any([cell.value == 'Средняя', cell.value == 'Держать']):
                cell.fill = PatternFill('solid', fgColor="ffff66")
            elif any([cell.value == 'Низкая', cell.value == 'Продавать', cell.value == 'Неликвидные']):
                cell.fill = PatternFill('solid', fgColor="FA8072")

    # Заливка ячеек. У автора есть доля в этой компании - зеленым, от компании лучше держаться подальше - красным
    my_positions = ['ALRS',
                    'KAZT',
                    'NVTK',
                    'GMKN',
                    'SBER',
                    'CHMF',
                    'TATN',
                    'TTLK',
                    'GAZP',
                    'LKOH',
                    'RASP',
                    ]
    dangerous_positions = ['KOGK',
                           ]
    for row in ws:
        for cell in row:
            for i in my_positions:
                if any([cell.value == i]):
                    cell.fill = PatternFill('solid', fgColor="32CD32")
            for j in dangerous_positions:
                if any([cell.value == j]):
                    cell.fill = PatternFill('solid', fgColor="FA8072")

    # Выравнивание влево для первых столбцов
    list_columns = ['A', 'B']
    for i in list_columns:
        for cell in ws[i]:
            cell.alignment = Alignment(wrapText=False, horizontal="left", vertical="center")

    # Заливка столбцов, используя свою функцию
    gr_yell_red(ws, 'No', 'I', 'L', 20, 10)  # Заливка столбцов с рентабельностью
    gr_yell_red(ws, 'Yes', 'M', 'M', 15, 15)  # Заливка столбца с рентабельностью P/E Шиллера
    gr_yell_red(ws, 'No', 'N', 'N', 40, 0)  # Заливка столбца Потенциал по Бенджамину Грэму, %
    gr_yell_red(ws, 'No', 'O', 'O', 40, 0)  # Заливка столбца Потенциал по Питеру Линчу, %
    gr_yell_red(ws, 'No', 'P', 'P', 40, 0)  # Заливка столбца Потенциал по прогнозируемому FCF, %

    # Закрепляем области, которые выше и левее указанной ячейки
    ws.freeze_panes = 'B2'

    # записываем файл
    wb.save(f'spreadsheets/{xl_name}')


today = datetime.datetime.today()

value_strategy(os.getenv('api_russian_stocks'), f'{today:%Y.%m.%d}_russian_stocks.xlsx')
value_strategy(os.getenv('api_china_stocks'), f'{today:%Y.%m.%d}_china_stocks.xlsx')

