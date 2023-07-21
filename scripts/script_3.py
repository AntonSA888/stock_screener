# pip install XlsxWriter pandas xlsxwriter xlrd xlwt openpyxl python-dotenv
import requests
import pandas as pd
from functions import *
import datetime
import openpyxl
import os
from dotenv import load_dotenv, find_dotenv


load_dotenv(find_dotenv())
today = datetime.datetime.today()
xl_name = f'{today:%Y.%m.%d}_us_growth_strategy.xlsx'


response = requests.get(os.getenv('api_us_growth_strategy'))
text = response.text
content = response.content

with open(f"responses/response_{xl_name}.txt", "w", encoding='utf8') as f:
    f.write(response.text)  # Записываем ответ от сервера в текстовый файл
    print('1. Ответ сервера записан в response...txt')

with open(f"responses/response_{xl_name}.txt", "r", encoding='utf8') as f:
    with open(f"responses/response_modify_{xl_name}.txt", "w", encoding='utf8') as m:
        for line in f:
            if not line.isspace():
                m.write(line)
        print('2. response...txt преобразован в response_modify...txt')

df_1 = pd.DataFrame({})  # Создаем пустой датафрейм, чтобы добавлять к нему строки
count = 0  # Счетчик строк в датафрейме
count_pass = 0  # Счетчик пропущенных строк

with open(f"responses/response_modify_{xl_name}.txt") as file:
    for line in file:
        list_ = list_float(line.removesuffix('\n').split('|'))
        # Проверка на количество пустых строк в списке-----------------------------
        # Пропустить список, если таких строк больше count_none
        count_none = 0
        for i in list_:
            try:
                if len(i) == 0:
                     count_none += 1
            except(TypeError):
                pass
        if count_none > 8:
            count_pass += 1
            continue
        # --------------------------------------------------------------------------
        ticker =              list_[0].removesuffix('.ME')
        last =                list_[1]
        date =                list_[2]
        piotroski =           list_[3]
        avg_roe_3y =          list_[4]
        roa_1y =              list_[5]
        net_margin_1y =       list_[6]
        cap =                 round(list_[7])

        df_1_new_row = pd.DataFrame({
            'Тикер': [ticker],
            'Цена': [last],
            'Капитализация, млн.': [cap],
            'Дата': [date],
            'Стратегия Роста США': [us_growth_stocks(piotroski, avg_roe_3y, roa_1y, net_margin_1y)],

        })

        df_1 = pd.concat([df_1, df_1_new_row], ignore_index=True)  # сцепить датафреймы
        count += 1

print(f'3. Создано строк в датафрейме: {count}')
print(f'4. Пропущено строк: {count_pass}')

# Указать writer библиотеки
writer = pd.ExcelWriter(f'spreadsheets/{xl_name}', engine='xlsxwriter')
df_1.to_excel(writer, 'Sheet1', index=False)  # Записать ваш DataFrame в файл

writer.close()  # Сохраним результат
print('5. Датафрейм записан в файл .xlsx')

# Наводим красоту --------------------------------------------------------------------------------
wb = openpyxl.load_workbook(f'spreadsheets/{xl_name}')  # читаем excel-файл
ws = wb['Sheet1']  # получаем лист, с которым будем работать
ws.row_dimensions[1].height = 75  # Высота первой строки в пикселах

# Задаем ширину столбцов НЕ в пикселах
ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 8
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 20
ws.column_dimensions['E'].width = 15

# Заливка и перенос текста во всём листе
for row in ws:
    for cell in row:
        cell.alignment = Alignment(wrapText=True, horizontal="center", vertical="center")  # Перенос текста
        if any([cell.value == 'Высокая', cell.value == '-', cell.value == 'Покупать']):
            cell.fill = PatternFill('solid', fgColor="32CD32")
        elif any([cell.value == 'Средняя', cell.value == 'Держать']):
            cell.fill = PatternFill('solid', fgColor="ffff66")
        elif any([cell.value == 'Низкая', cell.value == 'Продавать', cell.value == 'Неликвидные']):
            cell.fill = PatternFill('solid', fgColor="FA8072")

# Выравнивание влево для первых столбцов
list_columns = ['A', 'B']
for i in list_columns:
    for cell in ws[i]:
        cell.alignment = Alignment(wrapText=True, horizontal="left", vertical="center")

# Закрепляем области, которые выше и левее указанной ячейки
ws.freeze_panes = 'B2'

# записываем файл
wb.save(f'spreadsheets/{xl_name}')
