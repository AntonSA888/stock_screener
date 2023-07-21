import openpyxl
from openpyxl.styles import (
                        PatternFill, Border, Side,
                        Alignment, Font, GradientFill
                        )


def list_float(list_):  # Принимает список строк, и если нет ошибки, то строку превращает в число
    for i in range(0, len(list_)):
        try:
            list_[i] = float(list_[i])
        except ValueError:
            pass
    return list_


def if_none(x):
    try:
        if x/1:
            return x
    except TypeError:
        return '-'


def chestn(x):
    try:
        if x > -1.78:
            return 'Низкая'
        elif x > -2.22:
            return 'Средняя'
        elif x <= -2.22:
            return 'Высокая'
    except TypeError:
        return '-'


def effect(x):
    try:
        if x > 7:
            return 'Высокая'
        elif x > 5:
            return 'Средняя'
        elif x <= 5:
            return 'Низкая'
    except TypeError:
        return '-'


def ust_bankr(x):
    try:
        if x > 2.6:
            return 'Высокая'
        elif x > 1.8:
            return 'Средняя'
        elif x <= 1.8:
            return 'Низкая'
    except TypeError:
        return '-'


def ust_dolg(debt_equity_ratio, interest_cover, current_ratio):
    try:
        count = 0
        if debt_equity_ratio < 2:
            count += 1
        if interest_cover > 1:
            count += 1
        if current_ratio > 1:
            count += 1
        if count == 3:
            return 'Высокая'
        elif count == 2:
            return 'Средняя'
        elif count < 2:
            return 'Низкая'
    except TypeError:
        return '-'


def akc_stoim(chestn, ust_bankr, graham_potential):
    try:
        if all([chestn == 'Высокая', ust_bankr == 'Высокая', graham_potential > 40]):
            return 'Покупать'
        elif any([chestn == 'Низкая', ust_bankr == 'Низкая', graham_potential <= 0]):
            return 'Продавать'
        else:
            return 'Держать'
    except TypeError:
        return '-'


def gr_yell_red(ws, invert, start_col, finish_col, more_then, less_then):
    list_columns = [chr(i) for i in range(ord(start_col), ord(finish_col)+1)]
    for i in list_columns:
        for cell in ws[i]:
            try:
                if invert == 'No':
                    if cell.value > more_then:
                        cell.fill = PatternFill('solid', fgColor="32CD32")
                    elif cell.value > less_then:
                        cell.fill = PatternFill('solid', fgColor="ffff66")
                    elif cell.value <= less_then:
                        cell.fill = PatternFill('solid', fgColor="FA8072")
                if invert == 'Yes':
                    if cell.value > more_then:
                        cell.fill = PatternFill('solid', fgColor="FA8072")
                    elif cell.value > less_then:
                        cell.fill = PatternFill('solid', fgColor="ffff66")
                    elif cell.value <= less_then:
                        cell.fill = PatternFill('solid', fgColor="32CD32")
            except TypeError:
                pass

def likv(x):
    # свыше 600 млн. руб. Высоколиквидные
    # 60-600 - Среднеликвидные
    # 1-60 - Низколиквидные
    # менее 1 - Неликвидные
    try:
        if x < 1:
            return 'Неликвидные'
        elif 60 > x >= 1:
            return 'Низкая'
        elif 600 > x >= 60:
            return 'Средняя'
        elif x >= 600:
            return 'Высокая'
    except TypeError:
        return '-'

def us_growth_stocks(piotroski, avg_roe_3y, roa_1y, net_margin_1y):
    try:
        if all([piotroski > 7, avg_roe_3y >= 30, roa_1y >= 10, net_margin_1y >= 10]):
            return 'Покупать'
        elif all([5 < piotroski, 25 < avg_roe_3y, 8 < roa_1y, 8 < net_margin_1y]):
            return 'Держать'
        else:
            return 'Продавать'
    except TypeError:
        return '-'


